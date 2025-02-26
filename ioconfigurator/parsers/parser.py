"""
Парсинг данных из IO-листа.xlsx
Файл парсится пока не встретится пустая ячейка в столбце A

Для корректной работы в IO-листе должны содержаться столбцы со следующими названиями в любом порядке:
- № п/п
- Поз. обозн.
- Наименование сигнала
- Вид сигнала
- Граница диапазона нижняя
- Граница диапазона верхняя
- Единицы измерения
- Обозначение ПЛК
- № модуля п/п
- Тип модуля
- № канала
- Тип канала
- Тип сигнала
- Шкаф
- Vars
"""
import openpyxl
from abc import abstractmethod
from openpyxl.utils.exceptions import InvalidFileException
from pathlib import Path
from typing import Any

import ioconfigurator.parsers.errors as errors
import ioconfigurator.parsers.model as model

__all__ = ['ParserIO', 'ParserVars']


class ParserXlsx:
    def __init__(self, path: Path, **kwargs):
        try:
            self._wb = openpyxl.load_workbook(filename=path, data_only=True)
        except InvalidFileException:
            raise errors.WrongExtension(path.suffix)
        self._ws = self._wb.active
        if 'cols_name' in kwargs:
            self._first_row, self._columns = self._find_columns(kwargs['cols_name'])
        elif 'first_row' in kwargs and 'cols' in kwargs:
            self._first_row, self._columns = kwargs['first_row'], {x: x for x in kwargs['cols']}
        else:
            raise TypeError('абракадабра')

    def _check_merge_cells(self, x_row: int, y_col: int) -> (int, int):
        for r in self._ws.merged_cells.ranges:
            if r.bounds[0] <= y_col+1 <= r.bounds[2] and r.bounds[1] <= x_row <= r.bounds[3]:
                return r.bounds[1]-1, r.bounds[0]-1

        return x_row, y_col

    def _get_cell_value(self, cell: str | None, idx_r: int, idx_c: int) -> str:
        if cell is None:
            _x, _y = self._check_merge_cells(idx_r, idx_c)
            cell = self._ws.cell(_x + 1, _y + 1).value

            if cell is None:
                return ''

        return cell.strip()

    def _find_columns(self, cols: list[str]) -> (int, dict[str, int]):
        columns_dict = {col: -1 for col in cols}

        columns_to_find = len(columns_dict)
        for idx_r, row in enumerate(self._ws.iter_rows(values_only=True)):
            for idx_c, cell in enumerate(row):
                cell_value = self._get_cell_value(cell, idx_r, idx_c)

                if cell_value in columns_dict and columns_dict[cell_value] == -1:
                    columns_dict[cell_value] = idx_c
                    columns_to_find -= 1
            if columns_to_find == 0:
                return idx_r + 1, columns_dict

        raise errors.NotInit(self._ws.title, [key for key in columns_dict if columns_dict[key] == -1])

    @abstractmethod
    def _parse_line(self, *args, **kwargs):
        """
        парсинг строки
        """

    @abstractmethod
    def parse(self, *args, **kwargs):
        """
        парсинг самого файла
        """


C_N_SIGNAL = '№ п/п'
C_SYMBOL = 'Поз. обозн.'
C_NAME = 'Наименование сигнала'
C_TYPE_RAW_SIGNAL = 'Вид сигнала'
C_MIN_UNITS = 'Граница диапазона нижняя'
C_MAX_UNITS = 'Граница диапазона верхняя'
C_UNITS = 'Единицы измерения'
C_PLC = 'Обозначение ПЛК'
C_N_MODULE = '№ модуля п/п'
C_TYPE_MODULE = 'Тип модуля'
C_N_CHANNEL = '№ канала'
C_TYPE_CHANNEL = 'Тип канала'
# C_CONTACT = 'Тип контакта'
C_TYPE_SIGNAL = 'Тип сигнала'
C_BOX = 'Шкаф'


class ParserIO(ParserXlsx):
    def __init__(self, path: Path):
        super().__init__(
            path=path,
            cols_name=[C_N_SIGNAL, C_SYMBOL, C_NAME, C_TYPE_RAW_SIGNAL, C_MIN_UNITS, C_MAX_UNITS, C_UNITS,
                       C_PLC, C_N_MODULE, C_TYPE_MODULE, C_N_CHANNEL, C_TYPE_CHANNEL, C_TYPE_SIGNAL, C_BOX]
        )
        self._signal_numb = {
            'ai': -1,
            'ao': -1,
            'di': -1,
            'do': -1,
        }

    def _parse_line(self, idx: int, row: Any, exclude: list[str], ai: bool, ao: bool, di: bool, do: bool) -> model.ParsedLine | None:
        type_channel = self._get_cell_value(row[self._columns[C_TYPE_CHANNEL]], idx, self._columns[C_TYPE_CHANNEL])
        plc = self._get_cell_value(row[self._columns[C_PLC]], idx, self._columns[C_PLC])
        box = self._get_cell_value(row[self._columns[C_BOX]], idx, self._columns[C_BOX])

        try:
            self._signal_numb[type_channel.lower()] += 1
        except KeyError:
            print(type_channel.lower())

        if type_channel.lower() == 'ai' and not ai or \
                type_channel.lower() == 'ao' and not ao or \
                type_channel.lower() == 'di' and not di or \
                type_channel.lower() == 'do' and not do or \
                exclude is not None and f'{plc} {box}' in exclude:
            return None

        n_signal = self._get_cell_value(row[self._columns[C_N_SIGNAL]], idx, self._columns[C_N_SIGNAL])
        symbol = self._get_cell_value(row[self._columns[C_SYMBOL]], idx, self._columns[C_SYMBOL])
        name = self._get_cell_value(row[self._columns[C_NAME]], idx, self._columns[C_NAME])
        type_raw_signal = self._get_cell_value(row[self._columns[C_TYPE_RAW_SIGNAL]], idx,
                                               self._columns[C_TYPE_RAW_SIGNAL])
        min_units = self._get_cell_value(
            row[self._columns[C_MIN_UNITS]], idx, self._columns[C_MIN_UNITS]
        ).replace(',', '.')
        max_units = self._get_cell_value(
            row[self._columns[C_MAX_UNITS]], idx, self._columns[C_MAX_UNITS]
        ).replace(',', '.')
        units = self._get_cell_value(row[self._columns[C_UNITS]], idx, self._columns[C_UNITS])
        n_module = self._get_cell_value(row[self._columns[C_N_MODULE]], idx, self._columns[C_N_MODULE])
        type_module = self._get_cell_value(row[self._columns[C_TYPE_MODULE]], idx, self._columns[C_TYPE_MODULE])
        n_channel = self._get_cell_value(row[self._columns[C_N_CHANNEL]], idx, self._columns[C_N_CHANNEL])
        type_signal = self._get_cell_value(row[self._columns[C_TYPE_SIGNAL]], idx, self._columns[C_TYPE_SIGNAL])

        return model.ParsedLine(
            nLine=idx,
            nSignal=n_signal,
            nSignalRef=self._signal_numb[type_channel.lower()],
            symbol=symbol,
            name=name,
            typeRawSignal=type_raw_signal,
            min=min_units,
            max=max_units,
            units=units,
            plc=plc,
            nModule=n_module,
            typeModule=type_module,
            nChannel=n_channel,
            typeChannel=type_channel,
            typeSignal=type_signal,
            box=box,
            errors=None
        )

    def parse(self, exclude: list[str], ai: bool, ao: bool, di: bool, do: bool) -> list[model.ParsedLine]:
        parsed_lines: list[model.ParsedLine] = []

        for idxR, row in enumerate(self._ws.iter_rows(values_only=True)):
            if idxR < self._first_row:
                continue

            if row[0] is None:
                break

            if line := self._parse_line(idxR, row, exclude, ai, ao, di, do):
                parsed_lines.append(line)

        return parsed_lines


C_VAR_SYMBOL = 0
C_VARS = 1


class ParserVars(ParserXlsx):
    def __init__(self, path: Path):
        super().__init__(path=path, first_row=0, cols=[C_VAR_SYMBOL, C_VARS])

    def _parse_line(self, idx: int, row: Any) -> tuple[str, str] | None:
        symbol = self._get_cell_value(row[self._columns[C_VAR_SYMBOL]], idx, self._columns[C_VAR_SYMBOL])
        variable = self._get_cell_value(row[self._columns[C_VARS]], idx, self._columns[C_VARS])

        return symbol, variable

    def parse(self) -> dict[str, str]:
        parsed_vars: dict[str, str] = {}

        for idxR, row in enumerate(self._ws.iter_rows(values_only=True)):
            if idxR < self._first_row:
                continue

            if row[self._columns[C_VAR_SYMBOL]] is None or row[self._columns[C_VARS]] is None:
                break

            if line := self._parse_line(idxR, row):
                parsed_vars[line[0]] = line[1]

        return parsed_vars


if __name__ == '__main__':
    for i in ParserIO(Path('D:\\Python\\IOConfigurator\\_data\\IO-лист EPLAN.xlsx')).parse([], True, True, True, True):
        print(i)
